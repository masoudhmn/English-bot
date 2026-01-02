"""Excel file processing for word import"""

from typing import List, Tuple, Optional
from pathlib import Path
import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models import Word
from src.config import EXCEL_COLUMNS


async def process_excel_file(
    session: AsyncSession,
    file_path: Path,
    user_id: int
) -> Tuple[int, List[str]]:
    """
    Process Excel file and add words to database.
    
    Returns:
        Tuple of (number of words added, list of duplicate words)
    """
    # Load Excel file
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Get header row to map columns
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            header_lower = str(cell.value).lower().strip()
            headers[header_lower] = idx
    
    # Verify required columns exist
    required_columns = ["word", "definition"]
    for col in required_columns:
        if col not in headers:
            raise ValueError(f"Required column '{col}' not found in Excel file")
    
    added_count = 0
    duplicates = []
    
    # Process each row
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        # Extract data
        word_text = ws.cell(row_idx, headers["word"]).value
        definition_text = ws.cell(row_idx, headers["definition"]).value
        
        # Skip empty rows
        if not word_text or not definition_text:
            continue
        
        word_text = str(word_text).strip()
        definition_text = str(definition_text).strip()
        
        # Get optional fields
        example_text = None
        if "example" in headers:
            example_cell = ws.cell(row_idx, headers["example"]).value
            if example_cell:
                example_text = str(example_cell).strip()
        
        translation_text = None
        if "translation" in headers:
            translation_cell = ws.cell(row_idx, headers["translation"]).value
            if translation_cell:
                translation_text = str(translation_cell).strip()
        
        # Check if word already exists (case-insensitive)
        stmt = select(Word).where(Word.word.ilike(word_text))
        result = await session.execute(stmt)
        existing_word = result.scalar_one_or_none()
        
        if existing_word:
            duplicates.append(word_text)
            continue
        
        # Add new word
        new_word = Word(
            word=word_text,
            definition=definition_text,
            example=example_text,
            translation=translation_text,
            added_by=user_id
        )
        session.add(new_word)
        added_count += 1
    
    # Commit all changes
    await session.commit()
    
    return added_count, duplicates


def validate_excel_structure(file_path: Path) -> Tuple[bool, Optional[str]]:
    """
    Validate Excel file structure.
    
    Returns:
        Tuple of (is_valid, error_message)
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Check if file has at least header row
        if ws.max_row < 1:
            return False, "Excel file is empty"
        
        # Get headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip())
        
        # Check required columns
        if "word" not in headers:
            return False, "Missing required column: 'word'"
        if "definition" not in headers:
            return False, "Missing required column: 'definition'"
        
        return True, None
        
    except Exception as e:
        return False, f"Error reading Excel file: {str(e)}"


def create_sample_excel(file_path: Path) -> None:
    """
    Create a sample Excel template for word import.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Words"
    
    # Add headers
    headers = ["word", "definition", "example", "translation"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Add sample data
    sample_data = [
        ["serendipity", "The occurrence of events by chance in a happy way", 
         "Finding that old photo was pure serendipity", "اتفاق خوش‌شانسانه"],
        ["ephemeral", "Lasting for a very short time", 
         "The ephemeral beauty of cherry blossoms", "زودگذر، گذرا"],
        ["eloquent", "Fluent or persuasive in speaking or writing", 
         "She gave an eloquent speech", "شیوا، گویا"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row_idx, col_idx).value = value
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
